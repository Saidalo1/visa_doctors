import tablib
from django.db.models import Prefetch
from django.utils.translation import gettext_lazy as _
from import_export import resources, fields
from import_export.resources import ModelResource
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models import SurveySubmission, Question, Response, AnswerOption, InputFieldType


class SurveySubmissionResource(resources.ModelResource):
    """
    Resource for SurveySubmission model import/export.

    Dynamically creates fields for each question.
    For each question, a main field is added that aggregates responses.
    Additionally, for choice questions with hierarchical options,
    separate fields are created for each root option that has children.
    """
    id = fields.Field(column_name=_('ID'), attribute='id')
    status = fields.Field(column_name=_('Status'), attribute='status')
    created_at = fields.Field(column_name=_('Created At'), attribute='created_at')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Создаем кэш и набор иерархических вариантов для каждого вопроса
        self.hierarchical_options = {}
        # Получаем все вопросы, отсортированные по order
        questions = Question.objects.select_related('field_type').order_by('order')
        for question in questions:
            # Если вопрос с выбором, вычисляем иерархические варианты
            if question.input_type in ['single_choice', 'multiple_choice']:
                root_options = question.options.filter(parent__isnull=True).order_by('order', 'text')
                hierarchical_ids = set()
                for root_option in root_options:
                    # Если у корневого варианта есть потомки, то они будут в отдельной колонке
                    if root_option.get_descendants().exists():
                        descendant_ids = list(root_option.get_descendants().values_list('id', flat=True))
                        hierarchical_ids.update(descendant_ids)
                        hierarchical_ids.add(root_option.id)
                self.hierarchical_options[question.id] = hierarchical_ids

            # Основное поле для вопроса
            field_name = f"question_{question.id}"
            field_label = question.field_type.field_key if (
                        question.field_type and question.field_type.field_key) else question.title
            self.fields[field_name] = fields.Field(column_name=field_label, attribute=None)
            self.fields[field_name].question_id = question.id
            # Захватываем id вопроса через аргумент по умолчанию
            setattr(self, f"dehydrate_{field_name}",
                    lambda obj, qid=question.id: self._get_question_value(obj, qid))

            # Для вопросов с выбором с иерархическими опциями добавляем отдельные поля
            if question.input_type in ['single_choice', 'multiple_choice']:
                root_options = question.options.filter(parent__isnull=True).order_by('order', 'text')
                for root_option in root_options:
                    if root_option.get_descendants().exists():
                        sub_field_name = f"question_option_{question.id}_{root_option.id}"
                        sub_field_label = root_option.text
                        self.fields[sub_field_name] = fields.Field(column_name=sub_field_label, attribute=None)
                        self.fields[sub_field_name].question_id = question.id
                        self.fields[sub_field_name].root_option_id = root_option.id
                        setattr(self, f"dehydrate_{sub_field_name}",
                                lambda obj, qid=question.id, roid=root_option.id: self._get_question_option_value(obj,
                                                                                                                  qid,
                                                                                                                  roid))

    def get_queryset(self):
        """
        Optimize queryset for export with prefetching related data.
        """
        return SurveySubmission.objects.prefetch_related(
            Prefetch(
                'responses',
                queryset=Response.objects.select_related('question', 'question__field_type')
                .prefetch_related('selected_options')
            )
        )

    def get_export_order(self):
        """
        Define the order of fields in the export.
        """
        # Стандартные поля
        fields_order = ['id', 'status', 'created_at']
        # Добавляем поля вопросов и дополнительные колонки для иерархических опций
        questions = Question.objects.order_by('order')
        for question in questions:
            fields_order.append(f"question_{question.id}")
            if question.input_type in ['single_choice', 'multiple_choice']:
                root_options = question.options.filter(parent__isnull=True).order_by('order', 'text')
                for root_option in root_options:
                    if root_option.get_descendants().exists():
                        fields_order.append(f"question_option_{question.id}_{root_option.id}")
        return fields_order

    def _get_question_value(self, submission, question_id):
        """
        Return the aggregated answer for a question, excluding hierarchical options.
        Это значение включает только те ответы, которые не попали в отдельные колонки.
        """
        if submission.id not in self._cached_responses or question_id not in self._cached_responses[submission.id]:
            return ''
        response = self._cached_responses[submission.id][question_id]

        input_type = response.question.input_type
        field_type = response.question.field_type.field_type_choice if response.question.field_type else None

        if input_type == Question.InputType.TEXT:
            if field_type == InputFieldType.FieldTypeChoice.NUMBER:
                try:
                    return float(response.text_answer) if response.text_answer and response.text_answer.strip() else ''
                except (ValueError, TypeError):
                    pass
            return response.text_answer or ''

        selected_options = list(response.selected_options.all())
        if selected_options:
            # Для вопросов с выбором исключаем те опции, которые уже выводятся в отдельных колонках
            if input_type in ['single_choice', 'multiple_choice']:
                hierarchical_ids = self.hierarchical_options.get(response.question.id, set())
                non_hierarchical_options = [opt for opt in selected_options if opt.id not in hierarchical_ids]
            else:
                non_hierarchical_options = selected_options

            if non_hierarchical_options:
                option_texts = [opt.export_field_name if opt.export_field_name else opt.text for opt in
                                non_hierarchical_options]
                custom_option = [opt for opt in non_hierarchical_options if opt.has_custom_input]
                if custom_option and response.text_answer:
                    return f"{', '.join(option_texts)} - {response.text_answer}"
                return ', '.join(option_texts)

        return response.text_answer or ''

    def _get_question_option_value(self, submission, question_id, root_option_id):
        """
        Return the answer value for a specific root option group.
        Фильтруем выбранные опции, чтобы включить только те, что принадлежат заданной группе.
        """
        if submission.id not in self._cached_responses or question_id not in self._cached_responses[submission.id]:
            return ''
        response = self._cached_responses[submission.id][question_id]
        selected_options = list(response.selected_options.all())
        from app.models import AnswerOption
        # Получаем все descendant'ы для заданного корневого варианта
        descendant_ids = list(
            AnswerOption.objects.get(pk=root_option_id).get_descendants().values_list('id', flat=True))
        # Включаем и сам корневой вариант
        group_ids = set(descendant_ids + [root_option_id])
        filtered_options = [opt for opt in selected_options if opt.id in group_ids]
        option_texts = [opt.export_field_name if opt.export_field_name else opt.text for opt in filtered_options]
        if filtered_options:
            return ", ".join(option_texts)
        return ''

    # Оптимизация экспорта через кэширование ответов
    _cached_responses = {}

    def before_export(self, queryset, *args, **kwargs):
        """
        Override before_export to prefetch all responses and cache them for performance.
        """
        super().before_export(queryset, *args, **kwargs)
        submission_ids = list(queryset.values_list('id', flat=True))[:200]
        all_responses = Response.objects.filter(submission_id__in=submission_ids).select_related(
            'question', 'question__field_type'
        ).prefetch_related('selected_options')
        self._cached_responses = {}
        for response in all_responses:
            if response.submission_id not in self._cached_responses:
                self._cached_responses[response.submission_id] = {}
            self._cached_responses[response.submission_id][response.question_id] = response

    def get_export_headers(self):
        """
        Get the export headers for each field.
        """
        headers = []
        for field_name in self.get_export_order():
            field = self.fields.get(field_name)
            headers.append(field.column_name if field and field.column_name else field_name)
        return headers

    def after_export(self, queryset, dataset, *args, **kwargs):
        """
        Clean up after export.
        """
        self._cached_responses = {}
        return dataset

    def export_resource_fields(self, obj, fields):
        """
        Extract data from an object for export for the given fields.
        """
        row = []
        for field_name in fields:
            if field_name.startswith('question_'):
                method = getattr(self, f"dehydrate_{field_name}", None)
                if method:
                    value = method(obj)
                else:
                    value = ''
            else:
                field = self.fields.get(field_name)
                if field:
                    value = self.dehydrate_field(obj, field)
                else:
                    value = ''
            row.append(value)
        return row

    def dehydrate_field(self, obj, field):
        """
        Extract the value for a field from the object.
        """
        if hasattr(field, 'attribute') and field.attribute:
            value = field.get_value(obj)
        else:
            method_name = f"dehydrate_{field.column_name.lower()}"
            if hasattr(self, method_name):
                value = getattr(self, method_name)(obj)
            else:
                value = ''
        return value

    def export(self, queryset=None, *args, **kwargs):
        """
        Export data to a Dataset.
        """
        import time, io
        from openpyxl import load_workbook
        start_time = time.time()
        self.before_export(queryset, *args, **kwargs)
        if queryset is None:
            queryset = self.get_queryset()
        if kwargs.get('export_form'):
            queryset = self.filter_export(queryset, **kwargs)
        dataset = tablib.Dataset()
        headers = self.get_export_headers()
        dataset.headers = headers
        export_order = self.get_export_order()
        for obj in queryset:
            row = self.export_resource_fields(obj, export_order)
            dataset.append(row)
        file_format = kwargs.get('file_format', None)
        if file_format and hasattr(file_format, 'get_extension') and file_format.get_extension() == 'xlsx':
            try:
                xlsx_data = io.BytesIO()
                file_format.export_data(dataset, xlsx_data)
                xlsx_data.seek(0)
                wb = load_workbook(xlsx_data)
                ws = wb.active
                for idx, cell in enumerate(ws[1], 1):
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True)
                    column = get_column_letter(idx)
                    ws.column_dimensions[column].width = min(50, max(15, len(str(cell.value)) + 5))
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                formatted_dataset = tablib.Dataset()
                formatted_dataset.xlsx = output.getvalue()
                dataset = formatted_dataset
            except Exception as e:
                print(f"Error during Excel formatting: {str(e)}")
        end_time = time.time()
        self.after_export(queryset, dataset, *args, **kwargs)
        return dataset

    class Meta:
        model = SurveySubmission
        fields = ('id', 'status', 'created_at', 'updated_at')
        export_order = ('id', 'status', 'created_at', 'updated_at')
        use_export_filter = True


class ResponseResource(ModelResource):
    """Resource for Response model import/export."""

    class Meta:
        """Meta options for Response resource."""
        model = Response
        fields = 'id', 'submission__id', 'question__title', 'text_answer', 'created_at'
        export_order = 'id', 'submission__id', 'question__title', 'text_answer', 'created_at'


class QuestionResource(ModelResource):
    """Resource for Question model import/export."""

    field_type_choice_display = fields.Field(
        column_name=_('Field Type Choice Display'),
        attribute='get_field_type_choice_display'
    )
    input_type_display = fields.Field(
        column_name=_('Input Type Display'),
        attribute='get_input_type_display'
    )

    class Meta:
        """Meta options for Question resource."""
        model = Question
        fields = (
            'id', 'title', 'field_title', 'placeholder', 'is_required',
            'input_type', 'input_type_display', 'field_type_choice', 'field_type_choice_display',
            'order', 'field_type__title', 'created_at'
        )
        export_order = (
            'id', 'title', 'field_title', 'input_type', 'input_type_display',
            'field_type_choice', 'field_type_choice_display', 'field_type__title',
            'is_required', 'order', 'created_at'
        )


class AnswerOptionResource(ModelResource):
    """Resource for AnswerOption model import/export."""

    class Meta:
        """Meta options for AnswerOption resource."""
        model = AnswerOption
        fields = 'id', 'question__title', 'text', 'order', 'parent__text', 'is_selectable', 'has_custom_input'
        export_order = 'id', 'question__title', 'text', 'order', 'parent__text', 'is_selectable', 'has_custom_input'


class InputFieldTypeResource(ModelResource):
    """Resource for InputFieldType model import/export."""

    class Meta:
        """Meta options for InputFieldType resource."""
        model = InputFieldType
        fields = 'id', 'title', 'regex_pattern', 'error_message', 'description'
        export_order = 'id', 'title', 'regex_pattern', 'error_message', 'description'
